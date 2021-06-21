import openpyxl
import os


class Product:

    def __init__(self):
        self.type = None
        self.partNumber = None
        self.modelName = None
        self.series = None
        self.failure = {}

    def setInfo(self, attr):
        import os
        product = {'series': {}, 'model': {}}
        file = open('series.txt', 'r')
        for line in file:
            key, value = line.strip().split(':')
            product['series'][key.strip()] = value.strip()
        file.close()
        file = open('models.txt', 'r')
        for line in file:
            key, value = line.strip().split(':')
            product['model'][key.strip()] = value.strip()
        file.close()
        del os
        # print(attr['par-n'])
        self.type = attr['erp-b']
        if attr['erp-b'] == 'FLASH':
            self.partNumber = attr['par-n']
            self.series = product['series'][attr['pro-s']+attr['erp-f']]
            self.modelName = product['model'][attr['pro-m']]
        elif attr['erp-b'] == 'DRAM' or attr['erp-b'] == 'SERVER':
            self.partNumber = attr['par-n']
            self.series = attr['erpfa']
            self.modelName = attr['erp-p'].strip('W/')
        elif attr['erp-b'] == 'EP':
            self.partNumber = attr['par-n']
            self.series = attr['erpfa']
            self.modelName = attr['mspec']
        rc1 = attr['roca1']
        rc2 = attr['roca2']

        if rc2 is None or rc2 == 'NPF':
            self.failure[rc1] = self.series + '_' + self.modelName
        else:
            self.failure[rc1 + ' ' + rc2] = self.series + '_' + self.modelName

    def addFailure(self, attr):
        rc1 = attr['roca1']
        rc2 = attr['roca2']

        if rc2 is None or rc2 == 'NPF':
            self.failure[rc1] = self.series + '_' + self.modelName
        else:
            self.failure[rc1 + ' ' + rc2] = self.series + '_' + self.modelName


class Case:
    def __init__(self):
        self.number, self.member = None, None
        self.openDate, self.closedDate, self.status = None, None, None
        self.customer, self.endCustomer, self.product = None, None, Product()
        self.failure, self.localTAT, self.hqTAT = None, None, None


class Reader:
    def __init__(self):
        self.xls = self.xlSheet()
        self.col = self.columns()
        self.cas = self.cases()

    def cases(self):
        cases = []
        print('DEBUG::', 'start:', 2, 'end:', len(self.xls['A']), '\n')
        for i in range(2, len(self.xls['A']) + 1, 1):
            index = str(self.xls.cell(row=i, column=1).row)
            head = self.xls.cell(row=i, column=1).value
            if self.xls[self.col['Type'] + index].value == 'CFQR' and head is not None:
                case = Case()
                start = i
                end = None
                for j in range(start + 1, len(self.xls['A']) + 1, 1):
                    value2 = self.xls.cell(row=j, column=1).value
                    if value2 is '*':
                        end = self.xls.cell(row=j, column=1).row
                        break
                    end = len(self.xls['A'])
                if end is None:
                    end = len(self.xls['A'])+1
                # print('DEBUG::', 'Case Start:', 'A'+str(start), 'Case End:', 'A'+str(end))
                case.failure = []
                for k in range(start, end, 1):
                    index2 = str(k)
                    # print('DEBUG::', 'index2:', index2)
                    pn = self.xls[self.col['Innodisk PN'] + index2].value.split('-')[0][1]
                    if '-' not in self.xls[self.col['Innodisk PN'] + index2].value:
                        ctrlr = self.xls[self.col['Innodisk PN'] + index2].value
                    else:
                        ctrlr = self.xls[self.col['Innodisk PN'] + index2].value.split('-')[1][3:6]

                    info = {'erp-s': self.xls[self.col['ERP-Familes'] + index2].value,
                            'erpfa': self.xls[self.col['ERP-Familes'] + index2].value,
                            'pro-b': self.xls[self.col['Product BU'] + index2].value.title(),
                            'erp-p': self.xls[self.col['ERP-Product Line'] + index2].value,
                            'erp-b': self.xls[self.col['ERP-BU'] + index2].value,
                            'par-n': self.xls[self.col['Innodisk PN'] + index2].value,
                            'pro-m': self.xls[self.col['Innodisk PN'] + index2].value.split('-')[0][2:5],
                            'pro-s': pn + ctrlr,
                            'erp-f': self.xls[self.col['ERP-Flash'] + index2].value,
                            'mspec': self.xls[self.col['Model Spec'] + index2].value,
                            'roca1': self.xls[self.col['Root cause L1'] + index2].value,
                            'roca2': self.xls[self.col['Root cause L2'] + index2].value,
                            'loc-t': self.xls[self.col['Local FA TAT'] + index2].value,
                            'for-t': self.xls[self.col['Forward FA TAT'] + index2].value,
                            }
                    if start == int(index2):
                        case.number = self.xls[self.col['Doc#'] + index2].value.strip()
                        case.member = list(' '.join(self.xls[self.col['FAE'] + index2].value.split('_')).title().split('/'))
                        case.openDate = self.xls[self.col['Submit Date'] + index2].value
                        case.closedDate = self.xls[self.col['FA Date'] + index2].value
                        case.status = self.xls[self.col['Status'] + index2].value
                        ''' Customer Add'''
                        if self.xls[self.col['Customer Name'] + index2].value is None:
                            case.customer = self.xls[self.col['End Customer Name'] + index2].value.title()
                        else:
                            case.customer = self.xls[self.col['Customer Name'] + index2].value.title()
                        if self.xls[self.col['End Customer Name'] + index2].value is not None:
                            case.endCustomer = self.xls[self.col['End Customer Name'] + index2].value.title()
                        ''' Product Add '''
                        case.product.setInfo(info)
                        ''' Failure Add '''
                        rc1 = self.xls[self.col['Root cause L1'] + index2].value
                        rc2 = self.xls[self.col['Root cause L2'] + index2].value
                        if rc2 is None or rc2 == 'NPF':
                            if rc1 not in case.failure:
                                case.failure.append(rc1)
                        else:
                            if rc1+' '+rc2 not in case.failure:
                                case.failure.append(rc1+' '+rc2)
                    else:
                        ''' Failure Add '''
                        rc1 = self.xls[self.col['Root cause L1'] + index2].value
                        rc2 = self.xls[self.col['Root cause L2'] + index2].value
                        if rc2 is None or rc2 == 'NPF':
                            if rc1 not in case.failure:
                                case.failure.append(rc1)
                        else:
                            if rc1+' '+rc2 not in case.failure:
                                case.failure.append(rc1+' '+rc2)
                        case.product.addFailure(info)
                """ TAT Add"""
                case.localTAT = self.xls[self.col['Local FA TAT'] + index2].value
                case.hqTAT = self.xls[self.col['Forward FA TAT'] + index2].value
                # case.product.failure = case.failure
                cases.append(case)
            # pprint(vars(case), indent=4)
            # print()
        print()
        return cases

    def xlSheet(self):
        for root, dirs, files in os.walk("."):
            for file in files:
                if 'eFAE' in file:
                    return openpyxl.load_workbook(file).active
        return None

    def columns(self):
        field = {'*': None, 'Doc#': None, 'Type': None, 'Status': None, 'FAE': None, 'Submit Date': None,
                 'FA Date': None, 'Customer Name': None, 'FAE': None, 'Product BU': None,
                 'Innodisk PN': None, 'ERP-Familes': None, 'ERP-Flash': None, 'ERP-Product Line': None,
                 'Forward FA TAT': None, 'Local FA TAT': None, 'Root cause L1': None, 'Root cause L2': None,
                 'ERP-BU': None, 'End Customer Name': None, 'Model Spec': None}
        for i in range(1, 2, 1):
            for j in range(1, 145, 1):
                if self.xls.cell(row=i, column=j).value in field.keys():
                    field[self.xls.cell(row=i, column=j).value] = self.xls.cell(row=i, column=j).column_letter
        return field
