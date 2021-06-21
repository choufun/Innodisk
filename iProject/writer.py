class Point:
    def __init__(self, fa, date, local, hq):
        self.fa = fa
        self.date = date
        self.local = int(local)
        self.hq = int(hq)
        self.total = self.local + self.hq
        if self.total < 11:
            self.hit = 1
        else:
            self.hit = 0


class Writer:
    def __init__(self, cases):
        self.cases = cases
        self.export(self.cases)

    @staticmethod
    def export(cases):
        import openpyxl
        workbook = openpyxl.Workbook()
        sheet1 = workbook.create_sheet("Failures", 0)
        sheet2 = workbook.create_sheet("Workload", 1)
        # sheet1 = workbook.active
        sheet1.column_dimensions['A'].width = 15
        sheet1.column_dimensions['B'].width = 15
        sheet1.column_dimensions['C'].width = 15
        sheet1.column_dimensions['E'].width = 25
        sheet1.title = 'Failures'
        sheet1.append([])
        sheet1.append(['FLASH'])
        sheet1col = ['Case', 'Series', 'Model', 'Failures']
        sheet1.append(sheet1col)
        for attr in cases:
            for failure in attr.product.failure.keys():
                if attr.product.type == 'FLASH':
                    sheet1.append([attr.number, attr.product.series, attr.product.modelName, failure])
        sheet1.append([''])
        sheet1.append(['DRAM'])
        sheet1.append(sheet1col)
        for attr in cases:
            for failure in attr.product.failure.keys():
                if attr.product.type in ['DRAM', 'SERVER']:
                    sheet1.append([attr.number, attr.product.series, attr.product.modelName, failure])
        sheet1.append([''])
        sheet1.append(['EP'])
        sheet1.append(sheet1col)
        for attr in cases:
            for failure in attr.product.failure.keys():
                if attr.product.type == 'EP':
                    sheet1.append([attr.number, attr.product.series, attr.product.modelName, failure])
        sheet2.column_dimensions['A'].width = 15
        sheet2.column_dimensions['B'].width = 15
        sheet2.column_dimensions['C'].width = 8
        sheet2.column_dimensions['D'].width = 8
        sheet2.column_dimensions['E'].width = 8
        sheet2.column_dimensions['F'].width = 8
        sheet2.append(['Workload'])
        sheet2.append([])
        points = {}
        for case in cases:
            if case.closedDate is not None:
                if case.member[0] not in list(points.keys()):
                    points[case.member[0]] = []
                    points[case.member[0]].append(Point(case.number, case.openDate, case.localTAT, case.hqTAT))
                else:
                    points[case.member[0]].append(Point(case.number, case.openDate, case.localTAT, case.hqTAT))
        for member in points:
            sheet2.append([member])
            sheet2.append(['FA', 'Open Date', 'Local', 'HQ', 'Total', 'Hit'])
            for point in points[member]:
                sheet2.append([point.fa, point.date, point.local, point.hq, point.total, point.hit])
            sheet2.append([])
        workbook.save("FAE summary.xlsx")
        del openpyxl
