import openpyxl

from pprint import pprint

from datetime import datetime


class Writer:
    def __init__(self, stack):
        self.workbook = openpyxl.Workbook()
        self.s = stack

    def report(self):
        self.summary('Summary', 0)
        self.customers('Customers', 1)
        self.flash('FLASH', 2)
        self.dram('DRAM', 3)
        self.ep('EP', 4)
        self.save()

    def summary(self, title, page):
        sheet = self.workbook.create_sheet(title, page)
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 8
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 25
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 12
        sheet.append(['Year', 'Month', 'Case', 'FAE', 'Region', 'Type', 'Customer', 'BU', 'Status'])
        for case in self.s:
            sheet.append([case.y, case.m, case.n, case.fo, case.st, case.t, case.cu, case.bu, case.s])

    def customers(self, title, page):
        sheet = self.workbook.create_sheet(title, page)
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 8
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 10
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 35
        sheet.column_dimensions['I'].width = 35
        sheet.append(['Year', 'Month', 'Case', 'Region', 'Type', 'BU', 'Rank', 'Customer', 'End Customer'])
        for case in self.s:
            sheet.append([case.y, case.m, case.n, case.st, case.t, case.bu, case.cr, case.cu, case.ec])

    def flash(self, title, page):
        sheet = self.workbook.create_sheet(title, page)
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['H'].width = 25
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 10
        sheet.append(['Case', 'BU', 'Part Numbers', 'Interface', 'Series', 'Model', 'Root Cause 1', 'Root Cause 2',
                      'Flash', 'Year', 'Month'])
        for case in self.s:
            if case.bu == 'FLASH':
                for fa in case.fa:
                    sheet.append([case.n, case.bu, fa.pn, fa.ty, fa.se, fa.mn, fa.r1, fa.r2, case.ft, case.y, case.m])

    def dram(self, title, page):
        sheet = self.workbook.create_sheet(title, page)
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['H'].width = 30
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['J'].width = 10
        sheet.append(['Case', 'BU', 'Part Numbers', 'Series', 'Speed', 'Model', 'Root Cause 1', 'Root Cause 2',
                      'Year', 'Month'])
        for case in self.s:
            if case.bu == 'DRAM':
                for fa in case.fa:
                    sheet.append([case.n, case.bu, fa.pn, fa.se, fa.ty, fa.mn, fa.r1, fa.r2, case.y, case.m])

    def ep(self, title, page):
        sheet = self.workbook.create_sheet(title, page)
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 8
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 50
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['J'].width = 10
        sheet.append(['Case', 'BU', 'Part Numbers', 'Type', 'Series', 'Model', 'Root Cause 1', 'Root Cause 2',
                      'Year', 'Month'])
        for case in self.s:
            if case.bu == 'EP':
                for fa in case.fa:
                    sheet.append([case.n, case.bu, fa.pn, fa.ty, fa.se, fa.mn, fa.r1, fa.r2, case.y, case.m])

    def save(self):
        now = datetime.now()
        date_time = now.strftime("%m.%d.%Y - %H.%M.%S")
        new_filename = str("{} - FAE Report.xlsx".format(date_time))
        self.workbook.save(new_filename)
