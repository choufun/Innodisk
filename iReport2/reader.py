import openpyxl
from utility import *
from pprint import pprint
from case import *


class Reader:
    def __init__(self, filename):
        self.s = []
        self.e = openpyxl.load_workbook(filename).active
        self.c = {}

        for index, row in enumerate(self.e.iter_rows(min_row=1, max_row=1)):
            for cell_obj in row:
                self.c[cell_obj.value] = cell_obj.column_letter

    def lookahead(self, case, row):
        print("case index: {} {}".format(case.i, self.e['A' + str(case.i)].value))
        print("curr index: {} {}".format(row[0].row, row[0].value))
        print("next index: {} {}".format(row[0].row + 1, self.e['B' + str(row[0].row + 1)].value))
        print("last index: {} {}".format(len(self.e['B'])-1, self.e['B' + str(len(self.e['B']))].value))

        if case.i == len(self.e['B'])-1:
            self.s.append(case)
            print('last case added')
            debug(case)
            print('stack size:', len(self.s), '\n')
        else:
            if self.e['A' + str(row[0].row + 1)].value == '*':
                self.s.append(case)
                print('case added')
                debug(case)
                print('stack size:', len(self.s), '\n')
            else:
                case.add(self.e, self.c, row[0].row, 'u')
                print('...case updated...')
                print('stack size:', len(self.s), '\n')

    def scan(self):
        case = None
        for index, row in enumerate(self.e.iter_rows(min_row=2, max_row=len(self.e['B']))):
            if self.e[self.c["Type"] + str(row[0].row)].value != 'DOA':
                if row[0].value == "*":
                    if self.e[self.c["BU"]+str(row[0].row)].value == 'FLASH':
                        case = Flash(self.e, self.c, row[0].row)
                    elif self.e[self.c["BU"]+str(row[0].row)].value == 'DRAM':
                        case = DRAM(self.e, self.c, row[0].row)
                    elif self.e[self.c["BU"]+str(row[0].row)].value == 'EP':
                        case = EP(self.e, self.c, row[0].row)
                self.lookahead(case, row)

    def stack(self):
        return self.s


def debug(obj):
    if isinstance(obj, openpyxl.cell.cell.Cell):
        print("DEBUG::\tcolumn: {}".format(obj.column))
        print("DEBUG::\tcolumn_letter: {}".format(obj.column_letter))
        print("DEBUG::\tvalue: {}".format(obj.value))
        print("DEBUG::\trow: {}".format(obj.row))
        print("DEBUG::\tcol_idx: {}".format(obj.col_idx))
        print('\n\n')

    elif isinstance(obj, Flash) or isinstance(obj, DRAM) or isinstance(obj, EP):
        print("DEBUG:: {}".format(obj.spec()))
        for fa in obj.fa:
            pprint(fa.info(), indent=4)
