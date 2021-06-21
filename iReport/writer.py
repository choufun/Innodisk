import openpyxl
from openpyxl.styles import Alignment
from pprint import pprint


class Writer:
    def __init__(self, clerk):
        self.workbook = openpyxl.Workbook()
        self.caseSummary(clerk.stack)
        self.customerList(clerk.customerStack)
        self.caseByEngineer(clerk.annualStack)
        self.caseByClosure(clerk.closedStack)
        self.caseByFlashFailures(clerk.productStack)
        self.caseByDRAMFailures(clerk.productStack)
        self.caseByEPFailures(clerk.productStack)

        self.workbook.save('FAE Report.xlsx')

    def caseSummary(self, stack):
        sheet = self.workbook.create_sheet('Case Summary', 0)
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 50
        sheet.column_dimensions['F'].width = 50
        sheet.append(['Case', 'Customer', 'FAE', 'Series', 'Model', 'Failures'])

        for case in stack:
            sheet.append([case.number, case.customer, case.fae, case.series, case.model, ', '.join(case.failure)])

    def customerList(self, stack):
        sheet = self.workbook.create_sheet('Customer List', 1)
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 8
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 8
        sheet.column_dimensions['E'].width = 8
        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['G'].width = 30
        sheet.column_dimensions['H'].width = 35
        sheet.column_dimensions['I'].width = 50

        sheet.append(['Year', 'Month', 'Case', 'Type', 'Rank', 'Customer', 'End Customer', 'Application', 'Subject'])
        for year, months in stack.items():
            sheet.append([year])
            for month, cases in months.items():
                sheet.append(['', month])
                for case in cases:
                    sheet.append(['', '', case.number, case.type, case.rank, case.customer, case.endCustomer,
                                  case.application, case.subject])

    def caseByEngineer(self, stack):
        sheet = self.workbook.create_sheet('Engineer Cases', 2)
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 8
        sheet.column_dimensions['C'].width = 8
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 7
        sheet.column_dimensions['F'].width = 8
        sheet.column_dimensions['G'].width = 28
        sheet.column_dimensions['H'].width = 9
        sheet.column_dimensions['I'].width = 9
        sheet.column_dimensions['J'].width = 9
        sheet.column_dimensions['K'].width = 11
        sheet.column_dimensions['L'].width = 10

        sheet.append(['FAE', 'Year', 'Month', 'Case', 'Rank', 'BU', 'Customer', 'Local TAT', 'HQ TAT', 'Total TAT',
                      'In Progress', 'Case Total'])

        for engineer, years in stack.items():
            total = 0
            sheet.append([engineer])
            for year, months in years.items():
                for month, cases in months.items():
                    for case in cases:
                        if case.status == 'In progress':
                            sheet.append(['', year, month, case.number, case.rank, case.type, case.customer,
                                          case.localTAT, case.hqTAT, case.totalTAT, 1])
                        else:
                            sheet.append(['', year, month, case.number, case.rank, case.type, case.customer,
                                          case.localTAT, case.hqTAT, case.totalTAT])
                        total += 1
            sheet.append(['', '', '', '', '', '', '', '', '', '', '', total])
            sheet.append([])

    def caseByClosure(self, stack):
        sheet = self.workbook.create_sheet('Case Closures', 3)
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 8
        sheet.column_dimensions['C'].width = 8
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 7
        sheet.column_dimensions['F'].width = 8
        sheet.column_dimensions['G'].width = 27
        sheet.column_dimensions['H'].width = 9
        sheet.column_dimensions['I'].width = 9
        sheet.column_dimensions['J'].width = 9
        sheet.column_dimensions['K'].width = 9
        sheet.column_dimensions['L'].width = 10

        sheet.append(['FAE', 'Year', 'Month', 'Case', 'Rank', 'BU', 'Customer', 'Local TAT', 'HQ TAT', 'Total TAT',
                      'Hit Target', 'Case Total'])

        for engineer, years in stack.items():
            total = 0
            sheet.append([engineer])
            for year, months in years.items():
                for month, cases in months.items():
                    for case in cases:
                        sheet.append(['', year, month, case.number, case.rank, case.type, case.customer, case.localTAT,
                                      case.hqTAT, case.totalTAT, case.target])
                        total += 1
            sheet.append(['', '', '', '', '', '', '', '', '', '', '', total])
            sheet.append([])

    def caseByFlashFailures(self, stack):
        sheet = self.workbook.create_sheet('Flash Failures', 4)
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 8
        sheet.column_dimensions['C'].width = 8
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 15
        sheet.column_dimensions['J'].width = 28
        sheet.column_dimensions['K'].width = 30

        sheet.append(['BU', 'Year', 'Month', 'Interface', 'Case', 'Customer', 'Series', 'Model', 'Root Cause 1', 'Root Cause 2',
                      'Failure'])

        for BU, years in stack.items():
            if BU == 'Flash':
                total = 0
                sheet.append([BU])
                for year, months in years.items():
                    sheet.append(['', year])
                    for month, interfaces in months.items():
                        sheet.append(['', '', month])
                        for interface, cases in interfaces.items():
                            for case in cases:
                                for failure in case.failure:
                                    if len(failure.split()) > 1:
                                        sheet.append(['', '', '', interface, case.number, case.customer, case.series,
                                                      case.model, failure.split()[0], ''.join(failure.split()[1:]),
                                                      failure])
                                    else:
                                        sheet.append(['', '', '', interface, case.number, case.customer, case.series,
                                                      case.model, failure.split()[0], '', failure])
                                    total += 1

    def caseByDRAMFailures(self, stack):
        sheet = self.workbook.create_sheet('DRAM Failures', 5)
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 8
        sheet.column_dimensions['C'].width = 8
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 25
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 15
        sheet.column_dimensions['J'].width = 20

        sheet.append(['BU', 'Year', 'Month', 'Case', 'Customer', 'Series', 'Model', 'Root Cause 1', 'Root Cause 2',
                      'Failure'])

        for BU, years in stack.items():
            if BU == 'DRAM':
                total = 0
                sheet.append([BU])
                for year, months in years.items():
                    sheet.append(['', year])
                    for month, cases in months.items():
                        sheet.append(['', '', month])
                        for case in cases:
                            for failure in case.failure:
                                if len(failure.split()) > 1:
                                    sheet.append(['', '', '', case.number, case.customer, case.series, case.model,
                                                  failure.split()[0], ''.join(failure.split()[1:]), failure])
                                else:
                                    sheet.append(['', '', '', case.number, case.customer, case.series, case.model,
                                                  failure.split()[0], '', failure])
                                total += 1

    def caseByEPFailures(self, stack):
        sheet = self.workbook.create_sheet('EP Failures', 6)
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 8
        sheet.column_dimensions['C'].width = 8
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 18
        sheet.column_dimensions['F'].width = 18
        sheet.column_dimensions['G'].width = 50
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 15
        sheet.column_dimensions['J'].width = 20

        sheet.append(['BU', 'Year', 'Month', 'Case', 'Customer', 'Series', 'Model', 'Root Cause 1', 'Root Cause 2',
                      'Failure'])

        for BU, years in stack.items():
            if BU == 'EP':
                total = 0
                sheet.append([BU])
                for year, months in years.items():
                    sheet.append(['', year])
                    for month, cases in months.items():
                        sheet.append(['', '', month])
                        for case in cases:
                            for failure in case.failure:
                                if len(failure.split()) > 1:
                                    sheet.append(['', '', '', case.number, case.customer, case.series, case.model,
                                                  failure.split()[0], ''.join(failure.split()[1:]), failure])
                                else:
                                    sheet.append(['', '', '', case.number, case.customer, case.series, case.model,
                                                  failure.split()[0], '', failure])
                                total += 1
