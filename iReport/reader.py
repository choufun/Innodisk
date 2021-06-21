import openpyxl
import os
from case import *
from pprint import pprint


class Reader:
    def __init__(self):
        self.stack = []
        self.excel = None
        self.type = None
        self.case = None

        self.load()
        self.read()

    def load(self):
        for root, dirs, files in os.walk("."):
            for file in files:
                if "eFAE" in file:
                    self.excel = openpyxl.load_workbook(file).active

        for row in self.excel.iter_rows(min_row=1, max_row=1):
            for cell in row:
                # if cell.value == 'ERP-BU':
                if cell.value == 'Product BU':
                    self.type = cell.column_letter

    def read(self):
        for i, row in enumerate(self.excel.iter_rows(min_row=2, max_row=len(self.excel['A']))):
            # print('Row:', row)
            # print('Row Index:', row[0].row)

            index = row[0].row
            print(row)

            #if row[0].value == '*' and self.excel['D'+str(index)]:
            if row[0].value == '*':
                for cell in row:
                    if cell.column_letter == self.type:
                        if cell.value == 'FLASH':
                            print(index, '...new case found...', cell.value)
                            self.case = Flash(self.excel, str(index))
                            print(index, '...adding Flash info...')

                        elif cell.value == 'DRAM' or cell.value == "SERVER":
                            print(index, '...new case found...', cell.value)
                            self.case = DRAM(self.excel, str(index))
                            print(index, '...adding DRAM info...')

                        elif cell.value == 'EP':
                            print(index, '...new case found...', cell.value)
                            self.case = EP(self.excel, str(index))
                            print(index, '...adding EP info...')

                        if index == len(self.excel['A']):
                            self.stack.append(self.case)
                            print(index, '...last case added...')
                            print('Stack Size:', len(self.stack), '\n')

                        elif self.excel['A' + str(index+1)].value == '*':
                            self.stack.append(self.case)
                            self.case = None
                            print(index, '...case added...')
                            print('Stack Size:', len(self.stack), '\n')

            elif row[0].value is None:
                print(index, '...adding additional info...')
                self.case.addFailure(self.excel, str(index))

                if index == len(self.excel['A']):
                    self.stack.append(self.case)
                    print(index, '...last case added...')
                    print('Stack Size:', len(self.stack), '\n')

                elif self.excel['A' + str(index+1)].value == '*':
                    self.stack.append(self.case)
                    self.case = None
                    print(index, '...case added...')
                    print('Stack Size:', len(self.stack), '\n')
